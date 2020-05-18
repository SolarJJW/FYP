import os,datetime
import re, nltk, nltk.data
from string import digits

import openpyxl
import xlrd
from nltk import MWETokenizer, word_tokenize
from nltk.corpus import stopwords, PlaintextCorpusReader
from nltk.tokenize import WordPunctTokenizer
import json
import pymongo
from py2neo import Graph, Node, Relationship
#mongodb操作
client = pymongo.MongoClient(host='127.0.0.1',port=27017)
db = client.test
collections = db.essays


#实体
class t:
    title = ""
    author = ""
    address = {}
    keywords = ""

#两个词的分词
def getcizu(root):
    # 这里设置自己的文件夹
    wordlists = PlaintextCorpusReader(root, '.*')
    x = nltk.Text(wordlists.words('Superalloys_2008_21_30.txt'))
    # 自行修改自己所设置文件夹下txt的名字
    return x.collocations(20000)
    # 改动20000可以设置提取词组的数目

#分句工具
def splitSentence(paragraph):
    tokenizer = nltk.data.load('tokenizers/punkt/english.pickle')
    sentences = tokenizer.tokenize(paragraph)
    return sentences

#分词工具
def wordtokenizer(sentence):
    words = WordPunctTokenizer().tokenize(sentence)
    return words

#删除目录下所有文件
def del_file(path):
    ls = os.listdir(path)
    for i in ls:
        c_path = os.path.join(path, i)
        if os.path.isdir(c_path):
            del_file(c_path)
        else:
            os.remove(c_path)


if __name__ == '__main__':
    #时间戳
    now_time = datetime.datetime.now()
    #待处理路径
    path = "C:/Users/jjwjj/Desktop/毕业论文/Superalloy/superalloy_project/2-TXT转UDF-8"
    #输出路径
    path_out = "C:/Users/jjwjj/Desktop/毕业论文/Superalloy/superalloy_project/3-提取并删除基本信息和参考文献"

    #目录下所有文件
    files = os.listdir(path)
    leni = 0
    l = len(os.listdir(path))

    #对目录下每个文件进行处理
    for file in files:
        # mongo字典
        txt1 = {}
        leni = leni + 1
        if file[:1] == '~':
            leni = leni - 1
            l = l - 1
            continue
        #打印时间
        print(now_time)
        #构建文件输入输出目录
        filePath = path + '/' + file
        fileout = path_out + '/' + file

        # print('处理文件数 ', leni, ' / ', l)
        # print('目标路径: ', filePath)
        pa = os.path.split(filePath)

        # 去除空行-------------------------------------
        ttext = ""
        with open(os.path.abspath(filePath), 'r+', encoding='UTF-8') as fr, \
                open(os.path.abspath(fileout), 'w', encoding='UTF-8') as fd:
            for text in fr.readlines():
                if text.split():
                    ttext += text
            fr.seek(0)
            fr.truncate()
            fr.write(ttext)
            fr.seek(0)

            # 统计-----------------------------------------
            # 行数
            line_counts = 0
            # 词数
            word_counts = 0
            # 字数
            character_counts = 0

            for line in fr:
                words = line.split()
                word_counts += len(words)
                character_counts += len(line)
            print("单词个数：", word_counts)

            # 关键词获取-----------------------------------
            fr.seek(0)
            txt = fr.readlines()
            n = 0
            for line in txt:
                n = n + 1
                if "Keywords:" in line:
                    t.keywords = line.split(" ", 1)[1]
                    break
                if "Key words:" in line:
                    t.keywords = line.split(" ", 1)[1]
                    break
            # 文献基本信息
            fr.seek(0)
            k = 1
            for lines in fr:
                if k == 1:
                    t.title = lines.strip('\n')
                if k == 2:
                    t.author = lines.strip('\n').replace('1', '').replace('2', '').replace('3', '').replace('4', '')
                if k >= 3 and k <= n:
                    t.address[k] = lines.strip('\n')
                if (k > n + 1):
                    break
                k = k + 1
            # 标题
            print("标题：", t.title)
            txt1["title"] = t.title
            # 作者
            # t.author = t.author.encode("utf-8").translate(None, digits)
            print("作者：", t.author)
            txt1["author"] = t.author
            # 地址
            for i in range(3, n):
                if i == 3:
                    adr = "address" + str(i - 2)
                    txt1[adr] = t.address[i]
                    print("地址：", t.address[i])
                else:
                    adr = "address" + str(i - 2)
                    txt1[adr] = t.address[i]
                    print(t.address[i])
            # 关键词
            txt1["keywords"] = t.keywords
            print("关键词：", t.keywords)
            # 字数
            txt1["word_counts"] = str(word_counts)
            # 处理1 删除少于2个词的行，基本信息行，参考文献
            wordt = 0
            fr.seek(0)
            txt = fr.readlines()
            n1 = 1
            strr = ""
            for line in txt:
                if n1 < n + 1:
                    n1 = n1 + 1
                    continue
                if "References" in line:
                    break
                words = line.split()
                wordt = len(words)
                if wordt <= 2 and wordt >= 1:
                    n1 = n1 + 1
                    continue
                strr = strr + line
                n1 = n1 + 1
            fd.seek(0)
            fd.truncate()
            fd.seek(0)
            strr = strr[:-1]
            #
            # fr.seek(0)

            #去换行符
            strr.replace("\n","").replace("\r","")
            fd.write(strr)
            result = strr.split('. ')



            stop_words = set(stopwords.words('english'))

            filtered_sentence = []

            english_punctuations = [',', '.', ':', ';', '?', '(', ')', '[', ']', '!', '@', '#', '%', '$',
                                    '*', '/']  # 自定义英文表单符号列表
            sym = "[,.:;，?()[]!@#$%*]"
            for i in word_tokenize(strr):
                if i.lower() not in stop_words:  # 过滤停用词
                    if i not in english_punctuations:  # 过滤标点符号
                        filtered_sentence.append(i)
            article = " ".join(filtered_sentence)

            # print(article)

            keywords_dictionary = []
            data1 = xlrd.open_workbook('C:/Users/jjwjj/Desktop/毕业论文/article_keywords.xlsx')
            data2 = xlrd.open_workbook('C:/Users/jjwjj/Desktop/毕业论文/domain_phrase.xlsx')

            table1 = data1.sheet_by_index(0)
            table2 = data2.sheet_by_index(0)
            tem_keywords = []
            alloy_keywords = []
            # 第一个字典（专业知识）
            for i in range(1, 148):
                fenci = []

                for j in range(1, 17):
                    cell = str(table1.cell(i, j)).replace('\\xa0', ' ').strip('\n').strip('\r').strip('\t').strip('\'')
                    if (cell[:5] == "empty"):
                        continue
                    else:
                        replace = cell[:5]
                        cell = cell.replace(replace, '').strip('\'').strip(' ').strip()
                        fenci = []
                        alloy_keywords.append(cell)
                        for word in word_tokenize(cell):
                            fenci.append(word)

                        fencituple = tuple(fenci)
                        tem_keywords.append(fencituple)

            # 第二个字典（爬虫）
            for i in range(1, 9648):
                fenci = []
                cell = str(table2.cell(i, 2)).replace('\\xa0', ' ').replace('()', '').replace('[]', '').strip(
                    '\n').strip('\r').strip('\t').strip('\'')
                if (cell[:5] == "empty"):
                    continue
                else:
                    replace = cell[:5]
                    cell = cell.replace(replace, '').strip('\'').strip(' ').strip()
                    fenci = []
                    for word in word_tokenize(cell):
                        fenci.append(word)
                    fencituple = tuple(fenci)
                    tem_keywords.append(fencituple)

            # 第三个字典（长度为2词组）
            str3 = str(getcizu(path))
            listwords = str3.split(';')
            fenci = []
            test = []
            for word in listwords:
                txt = word.strip('[').strip(']').strip('\"')
                list = txt.split(',')
                for word in list:
                    word = word.strip("\'")
                    word = word.strip(" ")
                    word = word.strip("\'")
                    fenci = []
                    for tu in word_tokenize(word):
                        fenci.append(tu)
                    tuword = tuple(fenci)
                    tem_keywords.append(tuword)

            # print(tem_keywords)
            tokenizerkeywords = MWETokenizer(tem_keywords, separator=' ')
            dic = {}
            for word in tokenizerkeywords.tokenize(article.split()):
                if word not in dic:
                    dic[word] = 1
                else:
                    dic[word] += 1
            dic = sorted(dic.items(), key=lambda kv: (kv[1], kv[0]), reverse=True)

            result_path = os.path.join('C:/Users/jjwjj/Desktop/毕业论文/', "分词表.xlsx")
            print(result_path)
            print('开始写入 分词表.xlsx 文件 ' + result_path)
            if os.path.exists(result_path):
                print('分词表.xlsx已存在，在表后添加数据 ' + result_path)
                workbook = openpyxl.load_workbook(result_path)
            else:
                print('分词表.xlsx不存在，创建excel ' + result_path)
                workbook = openpyxl.Workbook()
                workbook.save(result_path)
            sheet = workbook.active
            headers = ["论文名", "单词", "词频"]
            # sheet.append(headers)
            numbers = ['1', '2', '3', '4', '5', '6', '7', '8', '9', '0']
            for k, v in dic:
                if v > 5 and len(k) > 2 and k[0] not in numbers or (len(word_tokenize(k)) >= 2 and len(k) > 3):
                    excel_row = []
                    excel_row.append(t.title)
                    excel_row.append(k)
                    excel_row.append(v)
                    # print(excel_row)
                    sheet.append(excel_row)
            workbook.save(result_path)
            print('生成 分词表.xlsx 文件 ' + result_path)


            result_path = os.path.join('C:/Users/jjwjj/Desktop/毕业论文/', "专业名词所在句子表.xlsx")
            print(result_path)
            print('开始写入 专业名词所在句子表.xlsx 文件 ' + result_path)
            if os.path.exists(result_path):
                print('分词表.xlsx已存在，在表后添加数据 ' + result_path)
                workbook = openpyxl.load_workbook(result_path)
            else:
                print('专业名词所在句子表.xlsx不存在，创建excel ' + result_path)
                workbook = openpyxl.Workbook()
                workbook.save(result_path)
            sheet = workbook.active
            headers = ["论文名", "单词", "句子"]
            # sheet.append(headers)
            # 逐行插入mongodb
            i = 1
            for se in result:
                s = "se" + str(i)
                se = se.strip()
                excel_row = []
                word_row = ""

                #se为句子，对此句子分词然后加到excel中
                for word in tokenizerkeywords.tokenize(se.split()):
                    if word in alloy_keywords:
                        word_row = word_row + word
                        word_row = word_row + ';'
                if word_row != "":
                    excel_row.append(t.title.strip(' '))
                    excel_row.append(word_row)
                    excel_row.append(se)
                    sheet.append(excel_row)
                    # print(excel_row)

                txt1[s] = se
                i = i + 1
            workbook.save(result_path)
            print('生成 专业名词所在句子表.xlsx 文件 ' + result_path)


        collections.insert_one(txt1)

    del_file(path)
    path = "C:/Users/jjwjj/Desktop/毕业论文/Superalloy/superalloy_project/1-PDF转TXT"
    del_file(path)
    path = "C:/Users/jjwjj/Desktop/毕业论文/Superalloy/superalloy_project/3-提取并删除基本信息和参考文献"
    del_file(path)
    path = "C:/Users/jjwjj/Desktop/毕业论文/Superalloy/superalloy_project/PDFs"
    del_file(path)